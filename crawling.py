from flask import Flask, render_template,request,url_for
from bs4 import BeautifulSoup
import requests
from selenium import webdriver
import sqlite3 as sql
from selenium.webdriver.common.keys import Keys
import time
from webdriver_manager.chrome import ChromeDriverManager
import csv
from openpyxl import Workbook,load_workbook

write_wb=Workbook()
ws1=write_wb.active
ws1.title='weather'
ws2=write_wb.create_sheet()
ws2.title='place'
ws3=write_wb.create_sheet()
ws3.title='corona'
ws4=write_wb.create_sheet()
ws4.title='safety'
ws5=write_wb.create_sheet()
ws5.title='vlog'

#날씨
def weather(city):
    browser=webdriver.Chrome(ChromeDriverManager().install())
    #browser=webdriver.Chrome('/Users/eunwoo/Travel_diary-master/chromedriver')
    browser.get("https://www.yahoo.com/news/weather")
    time.sleep(2)
    browser.find_element_by_css_selector("div.location-picker button.selector").click()
    browser.find_element_by_css_selector("input.search-input.Bd").send_keys(city+Keys.RETURN)
    time.sleep(2)

    resp=browser.page_source
    soup=BeautifulSoup(resp,'html.parser')
    browser.quit()

    day=[]
    pic=[]
    maxtemp=[]
    mintemp=[]

    today_wt=soup.select("div.forecast div.BdB div span")
    img=soup.select("div.forecast div.BdB span.Ta(c) img")
    min=soup.select("div.forecast div.BdB span span.low")
    max=soup.select("div.forecast div.BdB span span.high")

    for i in range(0,len(today_wt)):
        day.append(today_wt[i].text)
 
    for i in range(0,len(img)):
        pic.append(img[i]['src'])

    for i in range(0,len(min)):
        mintemp.append(min[i].text)
        mintemp[i]=mintemp[i][0:2]
        mintemp[i]=int((int(mintemp[i])-32)/1.8)

    for i in range(0,len(max)):
        maxtemp.append(max[i].text)
        maxtemp[i]=maxtemp[i][0:2]
        maxtemp[i]=int((int(maxtemp[i])-32)/1.8)

    for i in range(1,8):
        ws1.cell(i,1,day[i-1])
        ws1.cell(i,2,pic[i-1])
        ws1.cell(i,3,maxtemp[i-1])
        ws1.cell(i,4,mintemp[i-1])
    write_wb.save('/Users/eunwoo/Travel_diary-master/static/travel.xlsx')

    return day,pic,mintemp,maxtemp

#가볼만한곳
def place(city):
    browser=webdriver.Chrome(ChromeDriverManager().install())
    browser.get("https://search.naver.com/search.naver?sm=top_hty&fbm=1&ie=utf8&query="+city)
    resp=browser.page_source
    soup=BeautifulSoup(resp,'html.parser')
    browser.quit()

    global cntry
    cntry=soup.select_one("div.item span a")
    cntry=cntry.text.split(' ')[0]
    engcity=soup.select("div.title_area div span")[0].text

    ctimg=soup.select("div.spots ul li img")
    cityimg=[]

    for i in range(0,5):
        cityimg.append(ctimg[i]['src'])

        place=soup.select("div.spots ul li div.info b.name")
        placename=[]
        place_href=[]

    for i in range(0,5):
        placename.append(place[i].text)
        place_href.append("https://search.naver.com/search.naver?sm=top_hty&fbm=1&ie=utf8&query="+placename[-1])

    for i in range(1,6):
        ws2.cell(i,1,cityimg[i-1])
        ws2.cell(i,2,placename[i-1])
        ws2.cell(i,3,place_href[i-1])
    write_wb.save('/Users/eunwoo/Travel_diary-master/static/travel.xlsx')

    browser=webdriver.Chrome(ChromeDriverManager().install())
    browser.get("https://search.naver.com/search.naver?sm=top_hty&fbm=1&ie=utf8&query="+city+"여행 시차")
    resp=browser.page_source
    soup=BeautifulSoup(resp,'html.parser')
    browser.quit()

    timediff=soup.select("div.rel_answer_wrap div.inner span")[0].text
    if timediff=="차이없음":
        diff = 0
        return engcity,cntry,diff
    diff=timediff.split(' ')[0][:-2]
    print(timediff)
    plus = timediff.split(" ")[1]
    if plus =="느림":
        diff = "-"+diff
    else:
        diff = "+"+diff
    
    return engcity,cntry,diff

#코로나 정보
def corona():
    browser=webdriver.Chrome(ChromeDriverManager().install())
    browser.get("http://ncov.mohw.go.kr/bdBoardList_Real.do?brdId=1&brdGubun=14&ncvContSeq=&contSeq=&board_id=&gubun=")
    resp=browser.page_source
    soup=BeautifulSoup(resp,'html.parser')
    browser.quit()

    cor=soup.select("table.num tbody tr td.w_bold")
    covidcntry=[]
    covidnum=[]
    conum=[]
    cnum=[]

    for i in range(0,len(cor)-2):
        covidcntry.append(cor[i].text)

    cornum=soup.select("table.num tbody tr td")

    for i in range(1,len(cornum),2):
        cnum.append(cornum[i].text)

    for i  in range(0,len(covidcntry)):
        conum.append(cnum[i])

    for i in range(0,len(conum)):
        covidnum.append('\n\n'.join(conum[i].split()))

    covid_country=covidcntry.index(cntry)
    covid_confirmed=covidnum[covid_country]

    ws3.cell(1,1,covid_country)
    ws3.cell(2,1,covid_confirmed)

    write_wb.save('/Users/eunwoo/Travel_diary-master/static/travel.xlsx')

    return covid_country,covid_confirmed

#안전
def safe():
    # 외교부 해외안전정보 웹사이트 크롤링        
    browser=webdriver.Chrome(ChromeDriverManager().install())
    browser.get("https://www.0404.go.kr/dev/country.mofa?idx=&hash=&chkvalue=no2&stext=&group_idx=&alert_level=0")
    resp=browser.page_source
    soup=BeautifulSoup(resp,'html.parser')
    browser.quit()
    # 나라별 안전 정보 
    safe=soup.select("div.country_stage_box ul.country_list li")

    safelist=list(safe)    
    safety = ['' for i in range(len(safelist)) ]
    for i in range(0,len(safelist)):
        safety[i]=list(safelist[i])

    s=[]        
    for i in range(0,len(safety)):
        s.append([])
        for j in range(0,len(safety[i])):
            if '\n' in safety[i][j]:
                continue;
            else:
                s[i].append(safety[i][j])

    # html 태그 제거     
    for i in range(0,len(s)):
        for j in range(0,len(s[i])):
            if j==0: # 나라명
                s[i][j]=s[i][j].text
            else: # 안전정보
                s[i][j]=s[i][j].img['alt']

    cindex=[]
    for i in range(len(s)):
        if s[i][0]==cntry:
            for j in range(1,len(s[i])):
                cindex.append(s[i][j])

    for i in range(1,len(cindex)+1):
        ws4.cell(i,1,cindex[i-1])
    write_wb.save('/Users/eunwoo/Travel_diary-master/static/travel.xlsx')
    return cindex

#유튜브 브이로그 참고 영상
def vlog(city):
    browser=webdriver.Chrome(ChromeDriverManager().install())
    browser.get("https://www.youtube.com/results?search_query="+city+"+vlog")

    resp=browser.page_source
    soup=BeautifulSoup(resp,'html.parser')
    browser.quit()

    vlog_img_src=[]
    vlog_img=soup.select("div#contents div#contents div#dismissible ytd-thumbnail a#thumbnail img#img")

    for i in range(0,4):
        vlog_img_src.append(vlog_img[i]['src'])
    
    vlog_title=[]
    vlog_href=[]
    video_link=[]
    vlogtitle=soup.select("div#contents div#contents div#dismissible div.text-wrapper div#title-wrapper h3 a")
    
    for i in range(0,4):
        vlog_title.append(vlogtitle[i]['title'])
        vlog_href.append("https://www.youtube.com"+vlogtitle[i]['href'])
        video_link.append("https://www.youtube.com/embed/"+vlogtitle[i]['href'].replace("watch?v=",""))

    vlogchannel=soup.select("div#text-container yt-formatted-string a.yt-simple-endpoint")
    vlog_chname=[]
    vlog_ch_href=[]
    for i in range(0,10,2):
        vlog_chname.append(vlogchannel[i].text)
        vlog_ch_href.append("https://www.youtube.com"+vlogchannel[i]['href'])

    for i in range(1,5):
        ws5.cell(i,1,vlog_img_src[i-1])
        ws5.cell(i,2,vlog_title[i-1])
        ws5.cell(i,3,vlog_href[i-1])
        ws5.cell(i,4,vlog_chname[i-1])
        ws5.cell(i,5,vlog_ch_href[i-1])
        ws5.cell(i,6,video_link[i-1])
    write_wb.save('/Users/eunwoo/Travel_diary-master/static/travel.xlsx')

    #vlog_img_src
    #vlog_title
    #vlog_href
    #vlog_chname
    #vlog_ch_href
