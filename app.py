from flask import Flask, render_template,request,url_for
import requests
import sqlite3 as sql
import time
import crawling
import pandas as pd
from openpyxl import Workbook,load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)

@app.route('/')
def first_page():
    return render_template("first_page.html")

@app.route('/search',methods=['POST'])
def search():
    if request.method=='POST':
        city = request.form['city']

        crawling.weather(city)
        engcity,cntry,diff=crawling.place(city)
        crawling.corona()
        crawling.safe()
        crawling.vlog(city)

        load_wb = load_workbook("/Users/eunwoo/Travel_diary-master/static/travel.xlsx",data_only=True)

        weather_list=[]
        place_list=[]
        corona_list=[]
        safety_list=[]
        vlog_list=[]
        flag_list=[]
        

        ws_weather=load_wb['weather']
        for i in range(1,8):
            for j in range(1,5):
                weather_list.append([])
                weather_list[i-1].append(ws_weather.cell(i,j).value)

        ws_place=load_wb['place']
        for i in range(1,7):
            for j in range(1,4):
                place_list.append([])
                place_list[i-1].append(ws_place.cell(i,j).value)

        ws_corona=load_wb['corona']
        corona_list.append(ws_corona.cell(2,1).value)

        ws_safety=load_wb['safety']
        for i in range(1,ws_safety.max_row + 1):
            safety_list.append(ws_safety.cell(i,1).value)

        safety_list_len=len(safety_list)

        ws_vlog=load_wb['vlog']
        for i in range(1,4):
            for j in range(1,7):
                vlog_list.append([])
                vlog_list[i-1].append(ws_vlog.cell(i,j).value)
        
        load_wbflag = load_workbook("static/flag.xlsx",data_only=True)
        
        ws_wbflag=load_wbflag.worksheets[0]
        for i in range(1,656):
            for j in range(1,3):
                flag_list.append([])
                flag_list[i-1].append(ws_wbflag.cell(i,j).value)
        
        for i in range(0,655):
            if flag_list[i][0]==cntry:
                flagsrc=flag_list[i][1]
                break;

        return render_template('travel_info.html',city=city,engcity=engcity,cntry=cntry,diff = int(diff),
                                                weather_list=weather_list,
                                                place_list=place_list,
                                                corona_list=corona_list,
                                                safety_list=safety_list,
                                                safety_list_len=safety_list_len,
                                                vlog_list=vlog_list,
                                                flagsrc = flagsrc)

@app.route('/list')
def list():
    con = sql.connect("newdb.db")
    con.row_factory = sql.Row

    cur=con.cursor()
    cur.execute('''create table if not exists diarylist (date text, title text, diary text, photo BLOB)''')
    cur.execute("select * from diarylist")

    rows=cur.fetchall()

    #테이블 출력
    # cols = [column[0] for column in cur.description]
    # data_df = pd.DataFrame.from_records(data=rows, columns=cols)
    # print(data_df)

    return render_template("diary_list.html",rows=rows)

@app.route('/diarylist',methods=['POST','GET'])
def diary_list():
    if request.method =='POST':
        try:
            date=request.form['date']
            title=request.form['title']
            diary = request.form['diary']
            f=request.files['photofile']
            f.save("static/"+secure_filename(f.filename))
            #data=f.read('./uploads')
            
            with sql.connect("newdb.db")as con:
                cur=con.cursor()
                cur.execute("INSERT INTO diarylist(date,title,diary,photo) VALUES(?,?,?,?)",(date,title,diary,f.filename))
                con.commit()

        except:
            con.rollback()
        finally:
            return list()
            con.close()


@app.route('/mydiary')
def my_diary():
    return render_template("travel_diary.html")

@app.route('/diaryoutput/<int:diaryid>/')
def diary_output(diaryid):
    con = sql.connect("newdb.db")
    con.row_factory = sql.Row

    cur=con.cursor()
    cur.execute("select * from diarylist LIMIT 1 OFFSET "+str(diaryid))
    row=cur.fetchall()
    return render_template("diary_output.html",row = row[0])

if __name__=='__main__':
    app.run(host='0.0.0.0',debug=True)
